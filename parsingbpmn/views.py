from datetime import datetime
import re, openpyxl

from django.forms import model_to_dict
from django.http import HttpResponse
from django.shortcuts import render, redirect
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment

from .forms import ProcessForm, SystemForm, ContextualizationForm, ProfileForm, FusionForm, SelectContextForm
from .models import Process, Asset, System, Asset_has_attribute, Attribute, Asset_type, Attribute_value, \
    Threat_has_attribute, Threat_has_control, Context, Profile, Contextualization, profile_maturity_control, \
    Subcategory, Control, profile_has_subcategory, Subcategory_is_implemented_through_control, Category, \
    contextualization_has_maturity_levels, Maturity_level, Fusioncontext_has_context, Family, \
    Threat_has_family
from .bpmn_python_master.bpmn_python import bpmn_diagram_rep as diagram
from utils.fusion_functions import checkPriority, comparingmaturity, convertFromDatabase, convertToDatabase, createdict, profileupgrade,\
    comparingcontrols, fusionprofileandupgrade

# Create your views here.

def system_management(request):
    if request.method == 'POST':
        form = SystemForm(request.POST)
        if form.is_valid():
            form.save()
            last_system = System.objects.latest('id')
            return redirect('bpmn_process_management', last_system.pk)
    else:
        form = SystemForm()
    systems = System.objects.all()
    return render(request,'system_management.html',{
        'form':form,'systems':systems
    })

def bpmn_process_management(request,pk):
    if request.method == 'POST':
        form = ProcessForm(request.POST, request.FILES)
        if form.is_valid():
            saved_form = form.save(commit=False)
            saved_form.system_id = pk
            saved_form.save()
            last_process = Process.objects.latest('id')
            bpmn_graph = diagram.BpmnDiagramGraph()
            pk = last_process.pk
            bpmn_graph.load_diagram_from_xml_file(Process.objects.get(pk=pk).xml)
            lista = bpmn_graph.get_nodes()
            #print(lista)
            annotations=[]
            associations=[]

            for tuple in lista:
                for dizionario in tuple:
                    if type(dizionario) is dict:
                        try:
                            if dizionario['type'].endswith("textAnnotation"):
                                annotations.append(dizionario)
                        except KeyError:
                            print()
                        try:
                            if dizionario['type'].endswith("association"):
                                associations.append(dizionario)
                        except KeyError:
                            print()

            e=""
            for tuple in lista:
                for dizionario in tuple:
                    if type(dizionario) is dict:
                        if dizionario['type'].endswith("Task"):
                            attribute_value = []
                            id_task = dizionario['id']
                            x = dizionario["x"]
                            y = dizionario["y"]
                            width = dizionario["width"]
                            height = dizionario["height"]
                            position = x + ":" + y + ":" + width + ":" + height
                            if dizionario['type'].startswith("send"):
                                asset_type = Asset_type.objects.get(name="Send task")
                                e=""
                                for assoc in associations:
                                    if(id_task==assoc['association'][1]):
                                        target_ref=assoc['association'][2]
                                        for textAnn in annotations:
                                            if(target_ref==textAnn['id']):
                                                e= (textAnn['textAnnotation'][0][1])
                                    e = e.replace(" ","_")
                                    e = e.lower()
                                if e=="pec_communication":
                                    attribute_value.append(Attribute_value.objects.get(value="PEC communication"))
                                elif e=="mail_communication":
                                    attribute_value.append(Attribute_value.objects.get(value="Mail communication"))
                                elif e=="post_office_communication":
                                    attribute_value.append(Attribute_value.objects.get(value="Post office communication"))
                            elif dizionario['type'].startswith("receive"):
                                asset_type = Asset_type.objects.get(name="Receive task")
                                id_task = dizionario['id']
                                e=""
                                for assoc in associations:
                                    if(id_task==assoc['association'][1]):
                                        target_ref = assoc['association'][2]
                                        for textAnn in annotations:
                                            if (target_ref == textAnn['id']):
                                                e= (textAnn['textAnnotation'][0][1])
                                    e = e.replace(" ","_")
                                    e = e.lower()
                                if e == "pec_communication":
                                    attribute_value.append(Attribute_value.objects.get(value="PEC communication"))
                                elif e == "mail_communication":
                                    attribute_value.append(Attribute_value.objects.get(value="Mail communication"))
                                elif e == "post_office_communication":
                                    attribute_value.append(
                                        Attribute_value.objects.get(value="Post office communication"))
                            elif dizionario['type'].startswith("user"):
                                asset_type = Asset_type.objects.get(name="User task")
                                id_task = dizionario['id']
                                e=""
                                for assoc in associations:
                                    if(id_task==assoc['association'][1]):
                                        target_ref = assoc['association'][2]
                                        for textAnn in annotations:
                                            if (target_ref == textAnn['id']):
                                                e= (textAnn['textAnnotation'][0][1])
                                    e = e.replace(" ","_")
                                    e = e.lower()
                                if e=="online":
                                    attribute_value.append(Attribute_value.objects.get(value="Online"))
                                elif e=="offline":
                                    attribute_value.append(Attribute_value.objects.get(value="Offline"))
                            elif dizionario['type'].startswith("manual"):
                                asset_type = Asset_type.objects.get(name="Manual task")
                                attribute_value.append(Attribute_value.objects.get(value="Manual task"))
                            elif dizionario['type'].startswith("service"):
                                asset_type = Asset_type.objects.get(name="Service task")
                                id_task = dizionario['id']
                                e=""
                                for assoc in associations:
                                    if(id_task==assoc['association'][1]):
                                        target_ref = assoc['association'][2]
                                        for textAnn in annotations:
                                            if (target_ref == textAnn['id']):
                                                e= (textAnn['textAnnotation'][0][1])
                                    e = e.replace(" ","_")
                                e = e.lower()
                                if e=="statefull":
                                    attribute_value.append(Attribute_value.objects.get(value="Statefull"))
                                elif e=="stateless":
                                    attribute_value.append(Attribute_value.objects.get(value="Stateless"))
                            elif dizionario['type'].startswith("script"):
                                asset_type = Asset_type.objects.get(name="Script task")
                                attribute_value.append(Attribute_value.objects.get(value="Script task"))
                            elif dizionario['type'].startswith("business"):
                                asset_type = Asset_type.objects.get(name="Business rule task")
                                attribute_value.append(Attribute_value.objects.get(value="Business rule task"))
                            asset = Asset(name=dizionario['node_name'],bpmn_id=id_task,position=position, process=Process.objects.get(pk=pk),asset_type=asset_type)
                            asset.save()
                            attribute = []
                            for value in attribute_value:
                                attribute.append(Attribute.objects.get(asset_type=asset_type,attribute_value=value))
                            for a in attribute:
                                asset_has_attribute = Asset_has_attribute(asset=asset,attribute=a)
                                asset_has_attribute.save()
                        elif dizionario['type'].endswith("task"):
                            asset = Asset(name=dizionario['node_name'], process=Process.objects.get(pk=pk))
                            asset.save()

            return redirect('process_view_task_type', pk)
    else:
        form = ProcessForm()
    processes = Process.objects.filter(system=System.objects.get(pk=pk))
    check_box = []
    for process in processes:
        assets = Asset.objects.filter(process=process)
        check_attribute = False
        for asset in assets:
            if not Asset_has_attribute.objects.filter(asset=asset):
                check_attribute = True
        check_box.append(check_attribute)
    processes_info = zip(processes,check_box)
    return render(request,'bpmn_process_management.html',{
        'form':form,'processes_info':processes_info,'pk':pk, 'processes':processes
    })

def delete_system(request,pk):
    if request.method == 'POST':
        system = System.objects.get(pk=pk)
        system.delete()
    return redirect('system_management')

def delete_process(request,pk):
    if request.method == 'POST':
        process = Process.objects.get(pk=pk)
        system_id = process.system.pk
        process.delete()
    return redirect('bpmn_process_management',system_id)

def process_view_task_type(request,pk):
    task_list = Asset.objects.filter(process=Process.objects.get(pk=pk))
    check_attribute = False
    for task in task_list:
        if task.asset_type == None:
            check_attribute = True
    if check_attribute == True:
        asset_type = Asset_type.objects.all()
        system = Process.objects.get(pk=pk).system
        processes = Process.objects.filter(system=system)
        return render(request, 'process_view_task_type.html', {
            'task_list':task_list,'asset_type':asset_type,'pk':pk,'processes':processes
        })
    else:
        return redirect('process_view_attribute', pk)

def task_type_enrichment(request,pk):
    if request.method == "POST":
        assets_for_process = Asset.objects.filter(process=Process.objects.get(pk=pk))
        task_enrichment = []
        types = []
        for asset in assets_for_process:
            task_enrichment.append(request.POST.get(str(asset.pk)))
        for type in task_enrichment:
            if type != None:
                type = int(type)
                types.append(Asset_type.objects.get(pk=type))
            else:
                types.append(None)
        for asset,type in zip(assets_for_process,types):
            if type != None:
                x = Asset.objects.get(pk=asset.pk)
                x.asset_type = type
                x.save()
        return redirect('process_view_attribute',pk)
    else:
        return redirect('task_type_enrichment',pk)

def process_view_attribute(request,pk):
    task_list = Asset.objects.filter(process=Process.objects.get(pk=pk))
    check_attribute = False
    for task in task_list:
        if not Asset_has_attribute.objects.filter(asset=task):
            check_attribute = True
    if check_attribute==True:
        task_attributes = []
        list_attributes = []
        for task in task_list:
            task_attributes.append(Asset_has_attribute.objects.filter(asset=task))
        for attributes in task_attributes:
            if not attributes:
                list_attributes.append("empty")
            else:
                sub_list = []
                for element in attributes:
                    sub_list.append(element.attribute.attribute_value.value)
                list_attributes.append(sub_list)
        send = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Send task"))
        receive = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Receive task"))
        user = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="User task"))
        manual = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Manual task"))
        service = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Service task"))
        script = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Script task"))
        business = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Business rule task"))
        task_info = zip(task_list,list_attributes)
        system = Process.objects.get(pk=pk).system
        processes = Process.objects.filter(system=system)
        return render(request, 'process_view_attribute.html', {
                'task_info':task_info,'send':send,'receive':receive,'user':user,'manual':manual,'service':service,
                'script':script,'business':business,'pk':pk,'processes':processes})
    else:
        return redirect('threats_and_controls',pk)

def process_enrichment(request,pk):
    if request.method == "POST":
        task_list = Asset.objects.filter(process=Process.objects.get(pk=pk))
        pathfile=Process.objects.filter(id=pk)[0].xml

        check_attribute = False
        for task in task_list:
            if not Asset_has_attribute.objects.filter(asset=task):
                check_attribute = True
        if check_attribute == True:
            assets_for_process = Asset.objects.filter(process=Process.objects.get(pk=pk))
            attributes_enrichment = []
            attributes = []
            for asset in assets_for_process:
                attributes_enrichment.append(request.POST.get(str(asset.pk)))
            for attribute_enrichment in attributes_enrichment:
                if attribute_enrichment != None:
                    attribute_enrichment = int(attribute_enrichment)
                    attributes.append(Attribute.objects.get(pk=attribute_enrichment))
                else:
                    attributes.append(None)

            for asset,attribute in zip(assets_for_process,attributes):
                if attribute != None:
                    asset_has_attribute = Asset_has_attribute(asset=asset,attribute=attribute)

                    writeTextAnnotation_bpmn(pathfile,asset.position,asset.bpmn_id,attribute.attribute_value)


                    asset_has_attribute.save()

            return redirect('threats_and_controls',pk)
        else:
            assets_for_process = Asset.objects.filter(process=Process.objects.get(pk=pk))
            attributes_enrichment = []
            attributes = []
            for asset in assets_for_process:
                attributes_enrichment.append(request.POST.get(str(asset.pk)))
            for attribute_enrichment in attributes_enrichment:
                if attribute_enrichment != None:
                    attribute_enrichment = int(attribute_enrichment)
                    attributes.append(Attribute.objects.get(pk=attribute_enrichment))
                else:
                    attributes.append(None)

            for asset, attribute in zip(assets_for_process, attributes):
                if attribute != None:
                    Asset_has_attribute.objects.filter(asset=asset).update(attribute=attribute)
            return redirect('threats_and_controls', pk)
    else:
        return redirect('process_enrichment',pk)

import random
import string

def get_random_string(length):
    letters = string.ascii_letters
    result_str = ''.join(random.choice(letters) for i in range(length))
    return result_str

def writeTextAnnotation_bpmn(pathfile,position,taskId,attribute_value):
    textAnnotationId="TextAnnotation_"+get_random_string(7)
    textAnnotation="\
    <bpmn:textAnnotation id=\""+str(textAnnotationId)+"\">\n \
    <bpmn:text>"+str(attribute_value)+"</bpmn:text>\n \
    </bpmn:textAnnotation>\n"
    associationId="Association_"+get_random_string(7)
    association="<bpmn:association id=\""+str(associationId)+"\" sourceRef=\""+taskId+"\" targetRef=\""+textAnnotationId+"\" />"
    positionValues=position.split(":")
    x=positionValues[0]
    y=positionValues[1]
    width=positionValues[2]
    height=positionValues[3]

    stringToWrite=str(textAnnotation)+" "+str(association)
    shapetextAnn="<bpmndi:BPMNShape id=\""+textAnnotationId+"_di\" bpmnElement="+textAnnotationId+">\n\
        <dc:Bounds x=\""+str(int(x)+20)+"\" y=\""+str(int(y)+20)+"\" width=\""+str(width)+"\" height=\""+height+"\" />\n\
      </bpmndi:BPMNShape>\n"

    shapeAssoc="<bpmndi:BPMNShape id=\""+associationId+"_di\" bpmnElement="+associationId+">\n\
        <dc:Bounds x=\""+x+"\" y=\""+y+"\"/>\n\
        <dc:Bounds x=\""+str(int(x)+20)+"\" y=\""+str(int(y)+20)+"\"/>\n\
      </bpmndi:BPMNShape>\n"

    f = open(str(pathfile), "r+")
    stringFile=f.read()

    from xml.dom.minidom import parse, parseString

    datasource = open(str(pathfile)) #convert to minidom object
    minidomObject = parse(datasource)
    process = minidomObject.getElementsByTagName('bpmn:process')
    for e in process:
        if(taskId in e.toxml()):

            print("per il task "+str(taskId)+" il process è ")
            print(e.toxml())

    #task[0].firstChild.nodeValue

    #Process_1j43nxw


    #print(minidomObject.toxml()) #convert to xml string

def edit_process(request,pk):
    if request.method == "POST":
        assets = Asset.objects.filter(process=Process.objects.get(pk=pk))
        assets_type = []
        list_attributes = []
        for asset in assets:
            assets_type.append(asset.asset_type)
            list_attributes.append("empty")

        task_info = zip(assets,list_attributes)
        send = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Send task"))
        receive = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Receive task"))
        user = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="User task"))
        manual = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Manual task"))
        service = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Service task"))
        script = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Script task"))
        business = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Business rule task"))
        system = Process.objects.get(pk=pk).system
        processes = Process.objects.filter(system=system)
        return render(request, 'process_view_attribute.html', {
            'task_info': task_info, 'send': send, 'receive': receive, 'user': user, 'manual': manual,
            'service': service,'script': script, 'business': business, 'pk': pk, 'processes': processes})

def threats_and_controls(request,pk):
    process = Process.objects.get(pk=pk)
    assets = Asset.objects.filter(process=process)
    attributes = []
    threats = []
    controls = []
    for asset in assets:
        attributes.append(Asset_has_attribute.objects.filter(asset=asset))
    for list_attribute in attributes:
        for attribute in list_attribute:
            attribute = attribute.attribute
            threats.append(Threat_has_attribute.objects.filter(attribute=attribute))
    for threats_of_asset in threats:
        sublist_controls = []
        for threat in threats_of_asset:
            threat = threat.threat
            sublist_controls.append(Threat_has_family.objects.filter(threat=threat))
        controls.append(sublist_controls)

    # Threat has family va cambiata in threat has control, ossia andando ad associare i threat dell'enisa con i sottocontrolli CIS

    clear_list_threats = []
    for threat_list in threats:
        for threat in threat_list:
            if threat.threat not in clear_list_threats:
                clear_list_threats.append(threat.threat)


    clear_list_controls = []
    for control_of_asset in controls:
        for control_of_threat in control_of_asset:
            for control in control_of_threat:
                # dovrebbe essere control.control ma l'abbiamo adattato per le family
                if control.family not in clear_list_controls:
                    clear_list_controls.append(control.family)

    system = Process.objects.get(pk=pk).system
    processes = Process.objects.filter(system=system)
    return render(request, 'threats_and_controls.html', {
        'process_name':process.name,'clear_list_threats': clear_list_threats,'clear_list_controls':clear_list_controls,'pk':pk,'processes':processes
    })

def threat_modeling(request,pk):
    assets = Asset.objects.filter(process=Process.objects.get(pk=pk))
    attributes = []
    threats = []
    controls = []
    for asset in assets:
        attributes.append(Asset_has_attribute.objects.filter(asset=asset))
    for list_attribute in attributes:
        for attribute in list_attribute:
            attribute = attribute.attribute
            threats.append(Threat_has_attribute.objects.filter(attribute=attribute))
    for threats_of_asset in threats:
        sublist_controls = []
        for threat in threats_of_asset:
            threat = threat.threat
            #qui dovrebbe essere threat has control
            sublist_controls.append(Threat_has_family.objects.filter(threat=threat))
        controls.append(sublist_controls)

    controls_per_asset = []
    for asset in threats:
        list_controls = []
        for threat in asset:
            threat = threat.threat
            # qui dovrebbe essere threat has control
            controls_per_threat = Threat_has_family.objects.filter(threat=threat)
            for control in controls_per_threat:
                # qui dovrebbe essere contorl.control
                control= control.family
                if control not in list_controls:
                    list_controls.append(control)
        controls_per_asset.append(list_controls)

    threat_model_info = zip(assets, attributes, threats, controls,controls_per_asset)
    system = Process.objects.get(pk=pk).system
    processes = Process.objects.filter(system=system)
    return render(request, 'threat_modeling.html',{
        'threat_model_info':threat_model_info,'pk':pk,'processes':processes
    })

def export_threat_modeling(request,pk):
    if request.method == "POST":

        #help: https://djangotricks.blogspot.com/2019/02/how-to-export-data-to-xlsx-files.html
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-{name}-report.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'),
            name=Process.objects.get(pk=pk).name.replace(" ","_")
        )
        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Threat_modeling_REPORT'
        columns = ['Asset name', 'Asset type', 'Asset attributes', 'Threats','Policy per asset']
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = Font(name="Times New Roman",size=12,bold=True,color='FF0000')
            cell.border = Border(left=Side(border_style="thin",color='FF000000'),
                                 right=Side(border_style="thin",color='FF000000'),
                                 top=Side(border_style="thin",color='FF000000'),
                                 bottom=Side(border_style="thin",color='FF000000'),)

        assets = Asset.objects.filter(process=Process.objects.get(pk=pk))
        attributes = []
        threats = []
        for asset in assets:
            attributes.append(Asset_has_attribute.objects.filter(asset=asset))
        for list_attribute in attributes:
            for attribute in list_attribute:
                attribute = attribute.attribute
                threats.append(Threat_has_attribute.objects.filter(attribute=attribute))

        attributes_list = []
        for attribute in attributes:
            attr_sublist = []
            for element in attribute:
                attr_sublist.append(element.attribute.attribute_value.value)
            attributes_list.append(attr_sublist)

        threats_list = []
        for threat in threats:
            threat_sublist = []
            for element in threat:
                threat_sublist.append(element.threat.name)
            threats_list.append(threat_sublist)

        controls_per_asset = []
        for asset in threats:
            list_controls = []
            for threat in asset:
                threat = threat.threat
                controls_per_threat = Threat_has_control.objects.filter(threat=threat)
                for control in controls_per_threat:
                    control = control.control
                    if control not in list_controls:
                        list_controls.append(control)
            controls_per_asset.append(list_controls)

        for asset,attribute,threat,control in zip(assets,attributes_list,threats_list,controls_per_asset):
            row_num += 1

            if not threat:
                threat0 = ''
            else:
                threat0 = str(threat[0])

            # Define the data for each cell in the row
            row = [
                asset.name,
                asset.asset_type.name,
                str(attribute[0]),
                threat0,
                "CIS."+str(control[0].pk)+" - "+str(control[0])
            ]

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.font = Font(name="Times New Roman", size=11, bold=False, color='FF000000')
                cell.border = Border(left=Side(border_style="thin", color='FF000000'),
                                     right=Side(border_style="thin", color='FF000000'),
                                     top=Side(border_style="thin", color='FF000000'),
                                     bottom=Side(border_style="thin", color='FF000000'), )

            count_attr = 0
            old_row = row_num
            while count_attr < len(attribute)-1:
                count_attr += 1
                row_num += 1

                row = [
                    '',
                    '',
                    str(attribute[count_attr]),
                    ''
                ]

                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = cell_value
                    cell.font = Font(name="Times New Roman", size=11, bold=False, color='FF000000')
                    cell.border = Border(left=Side(border_style="thin", color='FF000000'),
                                         right=Side(border_style="thin", color='FF000000'),
                                         top=Side(border_style="thin", color='FF000000'),
                                         bottom=Side(border_style="thin", color='FF000000'), )

            count_threats = 0
            count_controls = 0
            row_num = old_row
            while count_threats < len(threat)-1 or count_controls < len(control)-1:
                row_num += 1

                if count_threats < len(threat)-1 and count_controls < len(control)-1:
                    count_threats += 1
                    count_controls += 1

                    row = [
                        '',
                        '',
                        '',
                        str(threat[count_threats]),
                        "CIS." + str(control[count_controls].pk) + " - " + str(control[count_controls])
                    ]
                elif count_threats < len(threat)-1 and not count_controls < len(control)-1:
                    count_threats += 1

                    row = [
                        '',
                        '',
                        '',
                        str(threat[count_threats]),
                        ''
                    ]
                else:
                    count_controls += 1

                    row = [
                        '',
                        '',
                        '',
                        '',
                        "CIS." + str(control[count_controls].pk) + " - " + str(control[count_controls])
                    ]

                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = cell_value
                    cell.font = Font(name="Times New Roman", size=11, bold=False, color='FF000000')
                    cell.border = Border(left=Side(border_style="thin", color='FF000000'),
                                         right=Side(border_style="thin", color='FF000000'),
                                         top=Side(border_style="thin", color='FF000000'),
                                         bottom=Side(border_style="thin", color='FF000000'), )
        #Per effettuare il resize delle celle in base a quella più grande
        dims = {}
        for row in worksheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            worksheet.column_dimensions[col].width = value

        workbook.save(response)

        return response

def bpmn_viewer(request,pk):
    process = Process.objects.get(pk=pk)
    return render(request,'bpmn_viewer.html',{
        'process':process
    })

def context_management(request):
    if request.method == 'POST':
        selectcontextform = SelectContextForm(request.POST)
        form = ContextualizationForm(request.POST)
        if form.is_valid():
            form.save()
            last_context = Context.objects.latest('id')
            return redirect('profile_management', last_context.pk)
    else:
        form = ContextualizationForm()
        selectcontextform = SelectContextForm(request.POST)
    context = Context.objects.all()
    return render(request,'context_management.html',{
        'form':form,'selectcontextform':selectcontextform, 'contexts':context
    })

def create_context(request):
    categories = Category.objects.all()
    subcategory_list=Subcategory.objects.all()
    form=ContextualizationForm(request.POST)
    priority_list=["Bassa", "Media", "Alta"]
    return render(request, 'create_context.html', {'form':form,'categories': categories, 'subcategory_list': subcategory_list, 'priority_list': priority_list, })

def save_contextualization(request):
    if request.method == 'POST':
        form=ContextualizationForm(request.POST)
        if form.is_valid():
            form.save()
            last_context=Context.objects.latest('id')

        subcategory_notclean = request.POST.getlist('subcategory')
        subcategory_list=[]
        for id in subcategory_notclean:
            subcategory_list.append(int(re.search(r'\d+', id).group()))

        priority= request.POST.getlist('priority')
        maturityname = request.POST.getlist('namemat')
        maturitydescription = request.POST.getlist('description')
        maturitylevel = request.POST.getlist('levelist')

        maturity_level = []

        for i, item in enumerate(maturityname, start=0):
            dict = {'name': item, 'description': maturitydescription[i], 'level': maturitylevel[i]}
            maturity_level.append(dict)

        for i,level in enumerate(maturity_level,start=0):
            newmaturitylevels=Maturity_level(name=level['name'], description=level['description'], level=level['level'], context_id=last_context.pk)
            newmaturitylevels.save()

        maturity_id = (Maturity_level.objects.filter(context_id=last_context.pk)).values()

        for i,subcategory in enumerate(subcategory_list, start=0):
            newcontextualization=Contextualization(subcategory_id=subcategory, context_id=last_context.pk, priority=priority[i])
            newcontextualization.save()
            for j,item in enumerate(maturity_id, start=0):
                newcontextualizationmaturity=contextualization_has_maturity_levels(maturity_level_id=item['id'], subcategory_contextualization_id=(Contextualization.objects.latest('id')).pk)
                newcontextualizationmaturity.save()

    fusionform=ContextualizationForm(request.POST)
    context = Context.objects.all()
    return render(request, 'context_management.html', {'contexts': context, 'form': fusionform})


def fusion_perform(request):
    if request.method == 'POST':
        selectform = SelectContextForm(request.POST)
        form= ContextualizationForm(request.POST)

        if form.is_valid() and selectform.is_valid():
            context_id1 = selectform['contextualization_1'].value()
            context_id2 = selectform['contextualization_2'].value()
            newcontext= []
            contextualization1 = Contextualization.objects.filter(context=Context.objects.get(pk=context_id1))
            contextualization2 = Contextualization.objects.filter(context=Context.objects.get(pk=context_id2))
            context1 = list(contextualization1.values())
            context2= list(contextualization2.values())

            context1= convertFromDatabase(context1)
            context2= convertFromDatabase(context2)
            i=0
            j=0

            while(i< len(context1) and j< len(context2)):
                if context1[i]['subcategory_id'] == context2[j]['subcategory_id']:

                    newelement = context1[i]

                    if(context1[i]['priority'] > context2[j]['priority']):
                        newelement['priority'] = context1[i]['priority']
                    else:
                        newelement['priority'] = context2[j]['priority']
                    if newelement['priority'] == 3:
                        checkPriority(context1[i]['maturity_level'], context2[j]['maturity_level'])

                    temp=[]
                    newelement['maturity_level']= comparingmaturity(context1[i]['maturity_level'], context2[j]['maturity_level'], temp)
                    newcontext.append(newelement)
                    i = i+1
                    j= j+1
                elif context1[i]['subcategory_id'] < context2[j]['subcategory_id']:
                    newelement = context1[i]
                    newcontext.append(newelement)
                    i=i+1
                else:
                    newelement =context2[j]
                    newcontext.append(newelement)
                    j=j+1

            while(i < len(context1)):
                newelement = context1[i]
                newcontext.append(newelement)
                i=i+1

            while(j< len(context2)):
                newelement=context2[j]
                newcontext.append(newelement)
                j=j+1

            newcontext= convertToDatabase(newcontext)
            form.save()
            last_context = Context.objects.latest('id')

            for item in newcontext:
                newcontextualization = Contextualization(context_id=last_context.pk, subcategory_id=item['subcategory_id'], priority= item['priority'])
                newcontextualization.save()

            lista_id=[]
            newmaturity= (Contextualization.objects.filter(context_id=last_context.pk)).values()
            for element in newmaturity:
                lista_id.append(element['id'])

            for i,item in enumerate(newcontext,start=0):
                for level in item['maturity_level']:
                    newcontexthasmaturity=contextualization_has_maturity_levels(subcategory_contextualization_id=lista_id[i], maturity_level_id=level)
                    newcontexthasmaturity.save()

            newfusioncontext = Fusioncontext_has_context(context_id=context_id1, fusion_context_id=last_context.pk)
            newfusioncontext.save()
            newfusioncontext2 = Fusioncontext_has_context(context_id=context_id2, fusion_context_id=last_context.pk)
            newfusioncontext2.save()

        return redirect('profile_management', last_context.pk)

    else:
        selectcontextform = SelectContextForm(request.POST)
        form = ContextualizationForm(request.POST)
    context = Context.objects.all()
    return render(request, 'context_management.html', {'context': context, 'selectcontextform ': selectcontextform , 'form': form })


def profile_management(request,pk):
    if request.method == 'POST':
        fusionform=FusionForm(request.POST)
        profileform=ProfileForm(request.POST)
    else:
        fusionform=FusionForm(request.POST)
        profileform = ProfileForm(request.POST)
    profiles = Profile.objects.filter(context=Context.objects.get(pk=pk))
    context=pk
    return render(request,'profile_management.html',{
        'profiles':profiles, 'fusionform': fusionform, 'context': context, 'profileform': profileform,
    })

def generate_profile(request, pk):
    if request.method == 'POST':
        profileform= ProfileForm(request.POST)
        if profileform.is_valid():
            saved_form = profileform.save(commit=False)
            saved_form.context_id = pk
            saved_form.save()
            last_profile= Profile.objects.latest('id')
            related_context= (Contextualization.objects.filter(context_id=pk)).values()
            chosenframework = last_profile.framework_id
            chosenlevel = (last_profile.level).lower()
            controls = list((Control.objects.filter(maturity_level=chosenlevel)).values())
            ids_controls=[]
            for control in controls:
                relatedframework= (Family.objects.filter(id=control['family_id'])).values('framework_id_id')
                if chosenframework == relatedframework[0]['framework_id_id']:
                    ids_controls.append(control)

            if chosenlevel.lower() == "minimo":
                chosenlevel = "1"
            elif chosenlevel.lower() == "standard":
                chosenlevel="2"
            else:
                chosenlevel="3"

            for contextrow in related_context:
                maturitylevels= (contextualization_has_maturity_levels.objects.filter(subcategory_contextualization_id=contextrow['id'])).values()
                for level in maturitylevels:
                    namelevel=(Maturity_level.objects.filter(id=level['maturity_level_id'])).values()
                    if chosenlevel in namelevel[0]['level'] :
                        profilelevel = namelevel[0]['id']
                profilesubcategory = profile_has_subcategory(profile_id=last_profile.pk, subcategory_id=contextrow['subcategory_id'], priority=contextrow['priority'], maturity_level_id=profilelevel)
                profilesubcategory.save()
                for control in ids_controls:
                    temp = (Subcategory_is_implemented_through_control.objects.filter(subcategory_id=contextrow['subcategory_id'], control_id=control['id'])).values()
                    if temp:
                        profilecontrols = profile_maturity_control(profile_id=last_profile.pk, subcategory_id=temp[0]['subcategory_id'], control_id=temp[0]['control_id'])
                        profilecontrols.save()

    profileform=ProfileForm(request.POST)
    fusionform= FusionForm(request.POST)
    profiles = Profile.objects.filter(context_id=pk)
    context = last_profile.context_id
    return render(request, 'profile_management.html', {'context':context ,'profiles': profiles, 'profileform':profileform, 'fusionform': fusionform})


def create_profile(request,pk):
    context= pk
    values_in_context=(Contextualization.objects.filter(context=context)).values()
    subcategory_dict=[]
    maturity_dict=[]
    priority_of_subcat=[]
    form= ProfileForm(request.POST)

    for value in values_in_context:
        temp=Subcategory.objects.filter(id=value['subcategory_id'])
        priority_of_subcat.append(value['priority'])
        subcategory_dict.append((list(temp.values()))[0])


    priority_list=["Bassa", "Media", "Alta"]
    request.session['list'] = subcategory_dict
    return render(request, 'create_profile.html', {'form':form, 'subcategory_dict': subcategory_dict, 'priority_list': priority_list,'priority_of_subcat': priority_of_subcat ,'context':context })

def save_profile(request,pk):
    subcategory_dict=request.session['list']
    if request.method == 'POST':
        form=ProfileForm(request.POST)
        if form.is_valid():
            saved_form = form.save(commit=False)
            saved_form.context_id = pk
            saved_form.save()
            last_profile= Profile.objects.latest('id')

        priority= request.POST.getlist('priority')


        sub_list=[]
        for dict in subcategory_dict:
            sub_id= dict['id']
            sub_list.append(sub_id)

        for i,subcategory in enumerate(sub_list,start=0):
            newprofilehassubcategory= profile_has_subcategory(profile_id=last_profile.pk, subcategory_id=subcategory, priority=priority[i])
            newprofilehassubcategory.save()

    request.session['list']=subcategory_dict
    return redirect('profile_controls', last_profile.pk)

def profile_controls(request,pk):
    subcategory_dict=request.session['list']
    profile = pk
    chosen_framework= (Profile.objects.filter(id=pk)).values('framework_id')
    controls_list = list((Control.objects.all()).values())
    ids_controls=[]
    subcategory_and_controls=[]

    for control in controls_list:
        relatedframework = (Family.objects.filter(id=control['family_id'])).values('framework_id_id')
        if chosen_framework[0]['framework_id'] == relatedframework[0]['framework_id_id']:
            ids_controls.append(control)

    for subcategory in subcategory_dict:
        temp=[]

        implemented = list((Subcategory_is_implemented_through_control.objects.filter(subcategory_id=subcategory['id'])).values())
        for control in ids_controls:
            i = 0
            while(i<len(implemented)):
                if control['id'] == implemented[i]['control_id']:
                    temp.append(control)
                i=i+1

        subcategory_and_controls.append({'subcategory': subcategory, 'related_controls': temp})

    context = (Profile.objects.get(pk=profile)).context_id
    request.session['subcategory_and_controls'] = subcategory_and_controls
    return render(request, 'profile_controls.html', {'subcategory_and_controls': subcategory_and_controls, 'ids_controls':ids_controls, 'profile': profile, 'context': context})

def save_profile_controls(request,pk):
    subcategory_and_controls = request.session.get('subcategory_and_controls')

    if request.method == 'POST':
        controls_notclean = request.POST.getlist('controls')
        implementation = request.POST.getlist('controls_implementation')
        clean_text=[]
        controls_and_implementation=[]

        for text in implementation:
            if (len(text) > 1):
                clean_text.append(text)

        for i,id in enumerate(controls_notclean,start=0):
            sub_and_control=id.split("\\")
            controls_and_implementation.append({'subcategory_id': sub_and_control[0], 'control': sub_and_control[1], 'implementation': clean_text[i]})

        for line in controls_and_implementation:
            profilecontro=profile_maturity_control(profile_id=pk, control_id=line['control'], subcategory_id=line['subcategory_id'], implementation=line['implementation'])
            profilecontro.save()

    profile = Profile.objects.get(pk=pk)
    context= profile.context_id
    return render(request, 'profile_controls.html', {'subcategory_and_controls': subcategory_and_controls, 'profile': profile.pk, 'context': context})

def profile_roadmap(request, pk):
    if request.method == 'POST':
        profile = profile_maturity_control.objects.filter(profile=Profile.objects.get(pk=pk))
        profilotarget = profile.values()
        dict_target = {}
        controls_target = createdict(profilotarget)
        implementedcontrols = controls_target
        implementation = "other"
        request.session['pk'] = pk
        request.session['implemented_list'] = implementedcontrols
        request.session['implementation'] = implementation
        return redirect('implemented_controls')
    else:
        return redirect('profile_management')

def implemented_controls(request):
    implementedcontrols =request.session['implemented_list']
    profilepk = request.session.get('pk')
    implementation=request.session.get('implementation')
    subcategory_clear_list = []
    controls_clear_list= []
    framework_clear_list=[]

    if implementation=="none":
        implementation_list = implementation
    else:
        implementation_list = (profile_maturity_control.objects.filter(profile_id=profilepk)).values()

    for control in implementedcontrols:
        element = control['subcategory_id']
        subcategory_clear_list.append((Subcategory.objects.get(pk=element)))
        framework_list=[]
        for i,element2 in enumerate(control['control_id'],start=0):
            control['control_id'][i] = Control.objects.get(pk=element2)
            family_id=(control['control_id'][i]).family_id
            frameworkid = (Family.objects.get(pk=family_id)).framework_id
            framework_list.append(frameworkid)
        controls_clear_list.append(control['control_id'])
        framework_clear_list.append(framework_list)

    return render(request, 'controls_implemented.html', {'profilepk': profilepk, 'implementation_list': implementation_list,'framework_clear_list': framework_clear_list, 'subcategory_clear_list': subcategory_clear_list, 'controls_clear_list': controls_clear_list})

def profile_evaluation(request,pk):
    if request.method == 'POST':
        profile_controls=(profile_maturity_control.objects.filter(profile=Profile.objects.get(pk=pk))).values()
        current_profile=Profile.objects.get(pk=pk)
        profile_framework=current_profile.framework_id
        context=current_profile.context_id
        profiles= (Profile.objects.filter(context_id=context)).values()
        min_profile = []
        std_profile = []
        avz_profile = []
        missing_controls=[]

        for profile in profiles:
            if "target" in profile['name']:
                if profile['framework_id'] == profile_framework:
                    if "minimo" in profile['name']:
                        min_temp = (profile_maturity_control.objects.filter(profile_id=profile['id'])).values()
                        min_profile = createdict(min_temp)
                    if "standard" in profile['name']:
                        std_temp= (profile_maturity_control.objects.filter(profile_id=profile['id'])).values()
                        std_profile=createdict(std_temp)
                    if "avanzato" in profile['name']:
                        avz_temp=(profile_maturity_control.objects.filter(profile_id=profile['id'])).values()
                        avz_profile = createdict(avz_temp)

        actual_profile = createdict(profile_controls)

        for subcat in actual_profile:
            for subcategory in min_profile:
                if subcategory['subcategory_id'] == subcat['subcategory_id']:
                    temp = []
                    newelement = comparingcontrols(subcat['control_id'],subcategory['control_id'], temp)
                    provamaturity = profile_has_subcategory.objects.filter(profile_id=current_profile, subcategory_id=subcat['subcategory_id'])
                    if newelement != []:
                        provamaturity.update(maturity_level_id=Maturity_level.objects.get(level=0,context_id=current_profile.context_id))
                        missing_controls.append({'subcategory_id': subcategory['subcategory_id'], 'control_id': newelement})
                    else:
                        provamaturity.update(maturity_level_id=Maturity_level.objects.get(level=1, context_id=current_profile.context_id))

        if not missing_controls:
            current_profile.level="minimo"
            for subcat in actual_profile:
                for subcategory in std_profile:
                    if subcategory['subcategory_id'] == subcat['subcategory_id']:
                        temp = []
                        newelement = comparingcontrols(subcat['control_id'], subcategory['control_id'], temp)
                        provamaturity = profile_has_subcategory.objects.filter(profile_id=current_profile,subcategory_id=subcat['subcategory_id'])
                        if newelement != []:
                            missing_controls.append({'subcategory_id': subcategory['subcategory_id'], 'control_id': newelement})
                        else:
                            matlev= Maturity_level.objects.get(level = 1, context_id=current_profile.context_id)
                            if provamaturity.values()[0]['maturity_level_id'] == matlev.id:
                                provamaturity.update(maturity_level_id=Maturity_level.objects.get(level=2, context_id=current_profile.context_id))

        else:
            current_profile.level="insufficiente"

        if not missing_controls:
            current_profile.level="standard"
            for subcat in actual_profile:
                for subcategory in avz_profile:
                    if subcategory['subcategory_id'] == subcat['subcategory_id']:
                        temp = []
                        newelement = comparingcontrols(subcat['control_id'], subcategory['control_id'], temp)
                        provamaturity = profile_has_subcategory.objects.filter(profile_id=current_profile, subcategory_id=subcat['subcategory_id'])
                        if newelement != []:
                            missing_controls.append({'subcategory_id': subcategory['subcategory_id'], 'control_id': newelement})
                        else:
                            matlev = Maturity_level.objects.get(level=2, context_id=current_profile.context_id)
                            if provamaturity.values()[0]['maturity_level_id'] == matlev.id:
                                provamaturity.update(maturity_level_id=Maturity_level.objects.get(level=3, context_id=current_profile.context_id))

        elif(str(current_profile.level) == "None"):
            current_profile.level="minimo"

        if not missing_controls:
            current_profile.level="avanzato"

        current_profile.save()

    request.session['missing_list'] = missing_controls
    return redirect('profile_management', context)

def profile_missing(request,pk):
    if request.method=='POST':
        profile=Profile.objects.get(pk=pk)
        context=profile.context_id
        missing_controls = request.session['missing_list']

        if (str(profile.level) == "None" or str(profile.level) == "avanzato"):
            missing_controls = []
        else:
            level=profile.level
            if level == "insufficiente":
                nextlevel="minimo"
            elif level=="minimo":
                nextlevel="standard"
            elif level=="standard":
                nextlevel="avanzato"

            nextprofile= Profile.objects.get(context_id=context, level=nextlevel)
            nextprofilelevel=(profile_maturity_control.objects.filter(profile_id=nextprofile.pk)).values()
            actualprofilelevel=(profile_maturity_control.objects.filter(profile_id=profile.pk)).values()

            controls_actual=createdict(actualprofilelevel)
            controls_target = createdict(nextprofilelevel)
            missing_controls = profileupgrade(controls_actual, controls_target)

        request.session['missing_list'] = missing_controls
        implementation= "none"
        request.session['implementation'] = implementation
        return redirect('controls_missing')
    else:
        return redirect('profile_management')

def controls_missing(request):
    missing_controls =request.session['missing_list']
    profilepk = request.session.get('pk')
    implementation=request.session.get('implementation')
    subcategory_clear_list = []
    controls_clear_list= []
    framework_clear_list=[]
    subcategory_actual_list = []
    subcategory_available = []

    if implementation=="none":
        implementation_list = implementation
    else:
        implementation_list = (profile_maturity_control.objects.filter(profile_id=profilepk)).values()

    for control in missing_controls:
        element = control['subcategory_id']
        subcategory_clear_list.append((Subcategory.objects.get(pk=element)))
        framework_list=[]
        for i,element2 in enumerate(control['control_id'],start=0):
            control['control_id'][i] = Control.objects.get(pk=element2)
            family_id=(control['control_id'][i]).family_id
            frameworkid = (Family.objects.get(pk=family_id)).framework_id
            framework_list.append(frameworkid)
        controls_clear_list.append(control['control_id'])
        framework_clear_list.append(framework_list)

    subcategory_actual = (profile_has_subcategory.objects.filter(profile_id=profilepk)).values()
    for subcategory in subcategory_actual:
        subcategory_actual_list.append(subcategory['subcategory_id'])

    for element in missing_controls:
        if element['subcategory_id'] in subcategory_actual_list:
            subcategory_available.append("A")
        else:
            subcategory_available.append("N/A")

    if request.method=='POST':
        request.session['missing_list']= missing_controls
    return render(request, 'controls_missing.html', {'subcategory_available': subcategory_available ,'profilepk': profilepk, 'implementation_list': implementation_list,'framework_clear_list': framework_clear_list, 'subcategory_clear_list': subcategory_clear_list, 'controls_clear_list': controls_clear_list})


def fusion_profile_perform(request):
    if request.method == 'POST':
        fusionform = FusionForm(request.POST)
        if fusionform.is_valid():
            actualprofile = fusionform['actual_profile'].value()
            targetprofile= Profile.objects.get(pk=fusionform['target_profile'].value())
            profiles = Profile.objects.filter(context_id=targetprofile.context_id)
            controls_minimo=[]
            controls_standard=[]
            controls_avanzato=[]

            for profile in profiles:
                if "minimo" in profile.name:
                    profileminimo = (profile_maturity_control.objects.filter(profile=profile.pk)).values()
                    controls_minimo = createdict(profileminimo)
                if "standard" in profile.name:
                    profilestandard = (profile_maturity_control.objects.filter(profile=profile.pk)).values()
                    controls_standard = createdict(profilestandard)
                if "avanzato" in profile.name:
                    profileavanzato = (profile_maturity_control.objects.filter(profile=profile.pk)).values()
                    controls_avanzato = createdict(profileavanzato)

            profileattuale= (profile_maturity_control.objects.filter(profile=Profile.objects.get(pk=actualprofile))).values()
            targetlevel = targetprofile.level
            controls_attuale = createdict(profileattuale)
            missingcontrols= fusionprofileandupgrade(controls_attuale, controls_minimo, controls_standard, controls_avanzato, targetlevel)

            request.session['missing_list']=missingcontrols

            request.session['implementation']="none"
            return redirect('controls_missing')
    else:
        fusionform = FusionForm(request.POST)
    profile= Profile.objects.all()
    return render(request,'profile_management.html',{
        'fusionform':fusionform,'profiles':profile
    })

def delete_context(request,pk):
    if request.method == 'POST':
        context = Context.objects.get(pk=pk)
        context.delete()
    return redirect('context_management')

def delete_profile(request,pk):
    if request.method == 'POST':
        profile = Profile.objects.get(pk=pk)
        context_id = profile.context.pk
        profile.delete()
    return redirect('profile_management', context_id)

def down_context_sample(request):
    with open('utils/Contestualizzazione_sample.xlsx', 'rb') as model_excel:
        result = model_excel.read()
    response = HttpResponse(result)
    response['Content-Disposition'] = 'attachment; filename=Contestualizzazione_sample.xlsx'
    return response

def read_context_file(request):
    if "GET" == request.method:
        return render(request, 'context_management.html')
    else:
        excel_file= request.FILES["excel_file"]
        wb=openpyxl.load_workbook(excel_file)
        worksheet=wb["Sheet1"]
        loadedcontext=Context(name="User Context")
        loadedcontext.save()
        last_context=(Context.objects.latest('id')).pk

        for i,row in enumerate(worksheet.iter_rows(),start=0):
            for cell in row:
                if cell.value == "X":
                    prioritylevel = row[4].value
                    subcategoryid = i
                    newcontextualization = Contextualization(priority=prioritylevel, subcategory_id=subcategoryid, context_id=last_context)
                    newcontextualization.save()
                    maturitylevelsraw= row[5].value
                    levels=maturitylevelsraw.split("\n")
                    for level in levels:
                        singlelevel=level.split(":")
                        condition=False
                        listmaturity=Maturity_level.objects.all()
                        for element in listmaturity:
                            if(singlelevel[0] == element.name):
                                if(singlelevel[1]== element.description):
                                    if(singlelevel[2]== element.level):
                                        if(last_context==element.context_id):
                                            maturityid=element.id
                                            condition=True

                        if(condition== False):
                            newmaturitylevel=Maturity_level(name=singlelevel[0], description=singlelevel[1], level=singlelevel[2], context_id=last_context)
                            newmaturitylevel.save()
                            maturityid=(Maturity_level.objects.latest('id')).pk
                        contextualization_id = Contextualization.objects.latest('id')
                        newcontextmaturity = contextualization_has_maturity_levels(maturity_level_id=maturityid,subcategory_contextualization_id=contextualization_id.pk)
                        newcontextmaturity.save()
        return redirect('profile_management', last_context)

def export_context(request, pk):
    if request.method == 'POST':
        contextualization_queryset=Contextualization.objects.filter(context_id=pk)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-{name}-context.xlsx'.format(
            date=datetime.now().strftime('%d-%m-%Y'),
            name=Context.objects.get(pk=pk).name.replace(" ","_")
        )
        workbook = Workbook()

        worksheet = workbook.active
        worksheet.title = 'Contextualization'
        columns = ['Function', 'Category', 'Subcategory', 'Priority_level', 'Maturity_level']
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = Font(name="Times New Roman", size=12, bold=True, color='FF0000')
            cell.border = Border(left=Side(border_style="thin", color='FF000000'),
                                 right=Side(border_style="thin", color='FF000000'),
                                 top=Side(border_style="thin", color='FF000000'),
                                 bottom=Side(border_style="thin", color='FF000000'), )

        # Iterate through all contextualization
        row_list = []

        for element in contextualization_queryset.values():
            subcategory= (Subcategory.objects.get(id=element['subcategory_id']))
            cat_id=(Subcategory.objects.get(id=element['subcategory_id']))
            category= (Category.objects.get(id=cat_id.category_id))
            priority_level= element['priority']
            contextualzation_maturity = (contextualization_has_maturity_levels.objects.filter(subcategory_contextualization_id=element['id'])).values()
            level = []
            for mat in contextualzation_maturity:
                mat_id=mat['maturity_level_id']
                temp=(Maturity_level.objects.get(id=mat_id))
                level.append(temp.name + ": " + temp.description)
            maturity_level= ';'.join(level)
            row_list.append({'Function': category.function, 'Category': category, 'subcategory': subcategory, 'priority_level': priority_level, 'maturity_level': maturity_level})

        for row in row_list:
            row_num +=1

            # define data for each cell in the row
            row = [
                row['Function'],
                row['Category'].name,
                row['subcategory'].name +": " + row['subcategory'].description,
                row['priority_level'],
                row['maturity_level']

            ]

            #assign data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.font = Font(name="Times New Roman", size=11, bold=False, color='FF000000')
                cell.border = Border(left=Side(border_style="thin", color='FF000000'),
                                     right=Side(border_style="thin", color='FF000000'),
                                     top=Side(border_style="thin", color='FF000000'),
                                     bottom=Side(border_style="thin", color='FF000000'), )

        for row in worksheet.rows:
            for cell in row:
                cell.alignment = Alignment(wrapText=True)
                value= cell.column_letter
                if(value != 'A' and value != 'D'):
                    worksheet.column_dimensions[value].width = 50
                else:
                    worksheet.column_dimensions[value].width = 20

        workbook.save(response)

    return response

def export_profile(request, pk):
    if request.method == 'POST':
        profile_queryset=profile_has_subcategory.objects.filter(profile_id=pk)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-{name}-profile.xlsx'.format(
            date=datetime.now().strftime('%d-%m-%Y'),
            name=Profile.objects.get(pk=pk).name.replace(" ","_")
        )
        workbook = Workbook()

        worksheet = workbook.active
        worksheet.title = 'Profile'
        columns = ['Function', 'Category', 'Subcategory', 'Priority_level', 'Maturity_level', 'Controls','Implementation']
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = Font(name="Times New Roman", size=12, bold=True, color='FF0000')
            cell.border = Border(left=Side(border_style="thin", color='FF000000'),
                                 right=Side(border_style="thin", color='FF000000'),
                                 top=Side(border_style="thin", color='FF000000'),
                                 bottom=Side(border_style="thin", color='FF000000'), )

        # Iterate
        row_list = []

        for element in profile_queryset.values():
            subcategory= (Subcategory.objects.get(id=element['subcategory_id']))
            cat_id=(Subcategory.objects.get(id=element['subcategory_id']))
            category= (Category.objects.get(id=cat_id.category_id))
            priority_level= element['priority']
            profile_maturity= (Maturity_level.objects.get(id=element['maturity_level_id']))
            maturity_level= profile_maturity.name + ": " + profile_maturity.description
            controls= (profile_maturity_control.objects.filter(profile_id=pk)).values()
            controlli=[]
            implementation=[]

            for control in controls:
                if subcategory.id == control['subcategory_id']:
                    controllo=Control.objects.get(id=control['control_id'])
                    controlli.append(controllo.name + ": " + controllo.description)
                    implementation.append(control['implementation'])

            controlsjoined = ";".join(controlli)
            if("target" in Profile.objects.get(pk=pk).name):
                controlimplementation="none"
            else:
                controlimplementation = ";".join(implementation)
            row_list.append({'Function': category.function, 'Category': category, 'subcategory': subcategory, 'priority_level': priority_level, 'maturity_level': maturity_level, 'controls':controlsjoined, 'implementation': controlimplementation })


        for row in row_list:
            row_num +=1

            # define data for each cell in the row
            row = [
                row['Function'],
                row['Category'].name,
                row['subcategory'].name +": " + row['subcategory'].description,
                row['priority_level'],
                row['maturity_level'],
                row['controls'],
                row['implementation']

            ]

            #assign data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.font = Font(name="Times New Roman", size=11, bold=False, color='FF000000')
                cell.border = Border(left=Side(border_style="thin", color='FF000000'),
                                     right=Side(border_style="thin", color='FF000000'),
                                     top=Side(border_style="thin", color='FF000000'),
                                     bottom=Side(border_style="thin", color='FF000000'), )
        for row in worksheet.rows:
            for cell in row:
                cell.alignment = Alignment(wrapText=True)
                value = cell.column_letter
                if (value != 'A' and value != 'D'):
                    worksheet.column_dimensions[value].width = 50
                else:
                    worksheet.column_dimensions[value].width = 20

        workbook.save(response)

    return response


def export_roadmap(request, pk):
    missingcontrols = request.session['missing_list']
    profilepk = request.session.get('pk')
    subcategory_actual_list=[]
    subcategory_available=[]

    subcategory_actual = (profile_has_subcategory.objects.filter(profile_id=profilepk)).values()
    for subcategory in subcategory_actual:
        subcategory_actual_list.append(subcategory['subcategory_id'])

    for element in missingcontrols:
        if element['subcategory_id'] in subcategory_actual_list:
            subcategory_available.append("A")
        else:
            subcategory_available.append("N/A")


    if request.method == 'POST':

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-roadmap.xlsx'.format(
            date=datetime.now().strftime('%d-%m-%Y'),

        )
        workbook = Workbook()

        worksheet = workbook.active
        worksheet.title = 'Roadmap'
        columns = ['Available','Subcategory', 'Controls']
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = Font(name="Times New Roman", size=12, bold=True, color='FF0000')
            cell.border = Border(left=Side(border_style="thin", color='FF000000'),
                                 right=Side(border_style="thin", color='FF000000'),
                                 top=Side(border_style="thin", color='FF000000'),
                                 bottom=Side(border_style="thin", color='FF000000'), )

        row_list = []

        for i,element in enumerate(missingcontrols,start=0):
            subcategory = (Subcategory.objects.get(id=element['subcategory_id']))
            controlli=[]

            for control in element['control_id']:
                controllo = Control.objects.get(id=control)
                controlli.append(controllo.name + ": " + controllo.description)
            controlsjoined = ";".join(controlli)

            row_list.append({'available': subcategory_available[i],'subcategory': subcategory, 'controls': controlsjoined})

        for row in row_list:
            row_num += 1

            # define data for each cell in the row
            row = [
                row['available'],
                row['subcategory'].name + ": " + row['subcategory'].description,
                row['controls']
            ]

            # assign data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.font = Font(name="Times New Roman", size=11, bold=False, color='FF000000')
                cell.border = Border(left=Side(border_style="thin", color='FF000000'),
                                     right=Side(border_style="thin", color='FF000000'),
                                     top=Side(border_style="thin", color='FF000000'),
                                     bottom=Side(border_style="thin", color='FF000000'), )

        for row in worksheet.rows:
            for cell in row:
                cell.alignment = Alignment(wrapText=True)
                value = cell.column_letter
                worksheet.column_dimensions[value].width = 50

        workbook.save(response)
    return response